import tkinter as tk
from tkinter import messagebox
import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import calendar
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import os
import time
import urllib3


def obter_caminhos(matricula):
    base_path = r"C:\Users\luisd\Downloads"
    return {
        "arquivo_origem": os.path.join(base_path, "todas-operacoes-diario.xlsx"),
        "arquivo_vigente": os.path.join(base_path, "todas-operacoes-viagentes.xlsx"),
        "arquivo_export": os.path.join(base_path, "export.xlsx"),
        "arquivo_destino": r"C:\Users\luisd\Downloads\base-balanço.xlsx"
    }

def confirmar_matricula():
    matricula = entry_matricula.get()
    email = entry_email.get()
    senha = entry_senha.get()
    if not matricula.isdigit():
        messagebox.showerror("Erro", "Número da matrícula deve conter apenas dígitos.")
        return
    if not email or not senha:
        messagebox.showerror("Erro", "Por favor, preencha o e-mail e a senha.")
        return

    if messagebox.askyesno("Confirmação", f"Confirma o número da matrícula: {matricula}?"):
        login_thunders(email, senha, matricula)

def executar_script(matricula):
    caminhos = obter_caminhos(matricula)

    try:
        # Função auxiliar para converter colunas de Excel em índices numéricos
        def coluna_para_indice(coluna):
            indice = 0
            for i, char in enumerate(reversed(coluna.upper())):
                indice += (ord(char) - ord('A') + 1) * (26 ** i)
            return indice - 1

        # Mapeamento das colunas de origem para destino
        colunas_mapeamento = {
            'K': 'A', 'D': 'B', 'F': 'C', 'G': 'D', 'R': 'E', 'Q': 'F',
            'P': 'G', 'O': 'H', 'AB': 'M', 'AA': 'N', 'AG': 'P',
            'T': 'S', 'AU': 'T', 'AV': 'U', 'AW': 'AC', 'AX': 'AD',
            'A': 'V',  
            'X': 'W',
            'V': 'X',
            'L': 'AB',
        }

        # Converter colunas de origem em índices numéricos
        indices_origem = {coluna_para_indice(origem): destino for origem, destino in colunas_mapeamento.items()}

        # Ler o arquivo de origem
        df_origem = pd.read_excel(caminhos["arquivo_origem"], header=None)

        # Converter a coluna de data para o tipo datetime
        indice_data_fornecimento = coluna_para_indice('K')
        df_origem[indice_data_fornecimento] = pd.to_datetime(
            df_origem[indice_data_fornecimento],
            format='%Y-%m-%d %H:%M:%S',
            errors='coerce'
        )

        # Definir o último mês e ano
        now = datetime.now()
        if now.month == 1:
            ultimo_mes = 12
            ano_corrente = now.year -1
        else:
            ultimo_mes = now.month -1
            ano_corrente = now.year

        # Filtrar dados do mês passado
        filtro_mes_passado = (
            (df_origem[indice_data_fornecimento].dt.month == ultimo_mes) &
            (df_origem[indice_data_fornecimento].dt.year == ano_corrente)
        )
        df_filtrado = df_origem[filtro_mes_passado]

        if df_filtrado.empty:
            messagebox.showerror("Erro", "Nenhum dado encontrado para o mês passado no arquivo de origem.")
            return

        # Criar DataFrame mapeado com as colunas desejadas
        colunas_destino = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + ["AA", "AB", "AC", "AD"]
        df_mapeado = pd.DataFrame(index=df_filtrado.index, columns=colunas_destino)
        for origem_idx, destino_col in indices_origem.items():
            df_mapeado[destino_col] = df_filtrado[origem_idx]

        # Formatar a coluna de data
        df_mapeado['A'] = pd.to_datetime(df_mapeado['A']).dt.strftime('%d/%m/%Y')

        # Ler o arquivo de vigentes
        df_vigente = pd.read_excel(caminhos["arquivo_vigente"], header=0)
        df_vigente['Período'] = pd.to_datetime(df_vigente['Período'], errors='coerce')

        # Filtrar dados do mês passado
        filtro_mes_passado_vigente = (
            (df_vigente['Período'].dt.month == ultimo_mes) &
            (df_vigente['Período'].dt.year == ano_corrente)
        )
        df_filtrado_vigente = df_vigente[filtro_mes_passado_vigente]

        if df_filtrado_vigente.empty:
            messagebox.showerror("Erro", "Nenhum dado encontrado para o mês passado no arquivo de vigentes.")
            return

        # Ajustar o tamanho do DataFrame filtrado
        df_filtrado_vigente = df_filtrado_vigente.head(len(df_mapeado))
        df_mapeado['I'] = df_filtrado_vigente['MWh Sazonalizado'].values

        # Calcular o número de horas no mês passado
        dias_no_mes_passado = calendar.monthrange(ano_corrente, ultimo_mes)[1]
        horas_no_mes_passado = dias_no_mes_passado * 24
        df_mapeado['AE'] = horas_no_mes_passado

        # Resetar o índice do DataFrame para garantir índices sequenciais
        df_mapeado.reset_index(drop=True, inplace=True)

        # Ler o arquivo exportado
        df_export = pd.read_excel(caminhos["arquivo_export"])

        if df_export.empty:
            messagebox.showerror("Erro", "O arquivo export.xlsx está vazio ou não foi encontrado.")
            return

        # Resetar o índice do DataFrame exportado
        df_export.reset_index(drop=True, inplace=True)

        # Carregar o workbook do Excel
        wb = load_workbook(caminhos["arquivo_destino"])

        # Definir as colunas que não devem ser sobrescritas
        columns_to_skip = ['J', 'K', 'L', 'O', 'Q', 'R', 'Z']
        indices_to_skip = [colunas_destino.index(col) for col in columns_to_skip]

        # Atualizar a aba "BASE" sem sobrescrever as fórmulas
        ws_base = wb["BASE"]
        for row_idx, row_data in df_mapeado.iterrows():
            for col_idx, value in enumerate(row_data):
                if col_idx in indices_to_skip:
                    continue  # Pula a escrita nesta coluna
                if pd.notnull(value):  # Verifica se o valor não é nulo
                    ws_base.cell(row=2 + row_idx, column=1 + col_idx, value=value)

        # Atualizar a aba "APOIO"
        ws_apoio = wb["APOIO"]
        for row_idx, row_data in df_export.iterrows():
            for col_idx, value in enumerate(row_data):
                if pd.notnull(value):  # Verifica se o valor não é nulo
                    ws_apoio.cell(row=1 + row_idx, column=1 + col_idx, value=value)

        # Salvar o workbook
        wb.save(caminhos["arquivo_destino"])
        messagebox.showinfo("Sucesso", "Dados copiados com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")

def acessar_url(navegador, url):
    """Abre a URL especificada no navegador."""
    navegador.get(url)
    time.sleep(2)

def preencher_campo(navegador, xpath, texto):
    """Preenche um campo localizado pelo XPath com o texto especificado."""
    campo = navegador.find_element(By.XPATH, xpath)
    campo.clear()
    campo.send_keys(texto)
    time.sleep(1)

def clicar_elemento(navegador, xpath):
    """Clica em um elemento localizado pelo XPath."""
    elemento = navegador.find_element(By.XPATH, xpath)
    elemento.click()
    time.sleep(1)

def login_thunders(email, senha, matricula):
    """Realiza login no sistema Thunders usando o e-mail e senha fornecidos."""
    try:
        # Configuração para o WebDriver não fechar automaticamente
        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach", True)

        # Ignorar verificação SSL ao instalar o driver do Chrome
        os.environ['WDM_SSL_VERIFY'] = '0'

        # Inicializar o WebDriver
        service = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(options=options, service=service)
        navegador.maximize_window()

        # Acessar o sistema e realizar login
        acessar_url(navegador, r"https://app.thunders.com.br/#/v2/home")
        preencher_campo(navegador, '//*[@id="Username"]', email)
        clicar_elemento(navegador, '/html/body/div/div/div[1]/form/button')
        preencher_campo(navegador, '//*[@id="Password"]', senha)
        clicar_elemento(navegador, '/html/body/div/div/div[1]/form/button')
        acessar_url(navegador, r"https://app.thunders.com.br/#/v2/operacoes/acl/todas")
        time.sleep(10)
        clicar_elemento(navegador, '//*[@id="exportar"]')
        clicar_elemento(navegador, '/html/body/app-root/ng-sidebar-container/div/div/div/app-acl/div/div/div/div/div/app-todas/div[2]/div[2]/div/div/button[1]')
        time.sleep(10)
        clicar_elemento(navegador, '//*[@id="exportar"]')
        clicar_elemento(navegador, '/html/body/app-root/ng-sidebar-container/div/div/div/app-acl/div/div/div/div/div/app-todas/div[2]/div[2]/div/div/button[2]')
        acessar_url(navegador, r"https://app.thunders.com.br/#/v2/operacoes/acl/vendas")
        time.sleep(10)
        clicar_elemento(navegador, '//*[@id="exportar"]')
        clicar_elemento(navegador, '/html/body/app-root/ng-sidebar-container/div/div/div/app-acl/div/div/div/div/div/app-vendas/div[2]/div[2]/div/div/button')
        
        executar_script(matricula)
        messagebox.showinfo("Sucesso", "Login realizado e arquivos acessados com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha no login ou navegação: {str(e)}")


# Interface gráfica
root = tk.Tk()
root.title("Login no Thunders e Confirmação de Matrícula")

tk.Label(root, text="Digite o número da matrícula:").pack(pady=5)
entry_matricula = tk.Entry(root, justify="center")
entry_matricula.pack(pady=5)

tk.Label(root, text="Digite o seu e-mail:").pack(pady=5)
entry_email = tk.Entry(root, justify="center")
entry_email.pack(pady=5)

tk.Label(root, text="Digite a sua senha do thunders:").pack(pady=5)
entry_senha = tk.Entry(root, justify="center", show="*")
entry_senha.pack(pady=5)

tk.Button(root, text="Confirmar e Fazer Login", command=confirmar_matricula).pack(pady=20)

root.mainloop()
